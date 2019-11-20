import math
import numpy as np
from scipy import integrate
from scipy.optimize import minimize
import scipy.stats as st
from matplotlib.patches import Ellipse
import matplotlib.pyplot as plt
import numdifftools as nd
from openpyxl import Workbook, load_workbook
from pyDOE import lhs
import copy
import logging
import time as time

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)



#Function measurement generates a measurement in silico given initial values for the states, time-invariant inputs, sampling time and standard deviation of measurement error
def transformation_ferraris(parameters):
    
    transformation = np.array([[-1.0, 1E4/(8.314*(273.15+105.0))], [0.0, 1.0]])
    
    new_parameters=np.dot(np.linalg.inv(transformation), parameters)
    
    return new_parameters, transformation


def measurement(x0,U,t,sigma,theta,model,measurable):
    
    X=integrate.odeint(model,x0,[0.0,t],(U, theta,))
    measurement=np.abs(X[1]+np.array(sigma)*np.random.normal(0.0,np.ones(len(measurable))))
    return [x0,U,t,measurement.tolist(),measurable]


#Function loglikelihood fits a model to the experimental data. It was preferred to integrate the model for every measurement singularly sacrifying computational efficiency

def loglikelihood(theta, dataset, sigma, model, measurable):
    
    residuals=np.zeros(shape=(len(dataset),len(sigma)))
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=[dataset[i][3] for i in range(0,len(dataset))]
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], theta))
        predictions[i]=X[1]
    
    residuals=((measurements-predictions)/sigma)**2
    
    objective=0
    
    for i in measurable:
        objective=objective+np.sum(residuals[:,i])
            
    return objective
      
      
#experiment performs a virtual experiment given initial states (list), time-invariant control input variables (list) and sampling times (list)  
  
def experiment(x0,U,st,sigma,theta,model,measurable):
    
    dataset=np.array([measurement(x0,U,st[0],sigma,theta,model,measurable)])
    for i in range(1,len(st)):
        dataset=np.append(dataset,[measurement(x0,U,st[i],sigma,theta,model,measurable)],axis=0)
        
    return dataset

#sensitivity evaluates the gradient of the measurable model output in the parameter space
def sensitivity(x0,U,t,theta,model,epsilon):
    #epsilon=0.0001
    sensitivity_matrix=np.ndarray(shape=(len(theta),len(x0)))
    perturbed_X=np.ndarray(shape=(len(theta),len(x0)))
    unperturbed_X=integrate.odeint(model,x0,[0.0,t],args=(U,theta,))[1]
    unperturbed_X=np.repeat([unperturbed_X],len(theta),axis=0)
    perturbed_theta=np.repeat([theta],len(theta),axis=0)+epsilon*np.eye(len(theta))*theta
    for i in range(0,len(theta)):
        perturbed_X[i]=integrate.odeint(model,x0,[0.0,t],args=(U,perturbed_theta[i],))[1]
    
    sensitivity_matrix=np.subtract(perturbed_X,unperturbed_X)/(epsilon*theta)
    
    return sensitivity_matrix
            

#expected Fisher calculates the Fisher information matrix related to a single measurement point
def expected_fisher(x0,U,t,theta,sigma,model,measurable):
    epsilon=0.0001
    Q=sensitivity(x0,U,t,theta,model,epsilon)
    Fisher=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        Fisher=Fisher+(1/(sigma[i]**2))*np.outer(Q[:,i],Q[:,i])
        
    return Fisher
    
    

        
def initial_guess_for_design(design_space, number_of_experiments, type_of_design=None):
    
    
    if type_of_design=='LatinHypercube':
        
        sampling=lhs(len(design_space), samples=number_of_experiments)
        
    elif type_of_design=='Uniform':
        
        sampling=np.random.uniform(0.0,1.0,size=len(design_space)*number_of_experiments).reshape(number_of_experiments, len(design_space))
    
    else:
        
        print 'The type of design selected is not available'
        
        return
        
    
    design=np.array([np.add(np.multiply(np.subtract(design_space[:,1], design_space[:,0]), sampling[i]), design_space[:,0]) for i in range(len(sampling))])
    
    
    return design
            
            
       
def exp_design(design_vector, model, length, measurable, parameters, observed_information, sigma, no_exp, criterion, output):
    
    conditions=design_vector.reshape((no_exp,3))
    
    F_hat=np.zeros(shape=(observed_information.shape))
    
    
    
    F_hat=sum(expected_fisher([conditions[i,0],0.0],[conditions[i,1],conditions[i,2]],length,parameters,sigma, model, measurable) for i in range(0,no_exp))
    
    if output=='objective':
    
        if criterion=='D':
        
            objective=np.log(np.linalg.det(np.linalg.inv(F_hat+observed_information)))
        
        elif criterion=='A':
        
            objective=np.trace(np.linalg.inv(F_hat+observed_information))
        
        elif criterion=='E':
        
            e,v=np.linalg.eig(np.linalg.inv(F_hat+observed_information))
            objective=np.max(e)
            
        elif criterion=='SV':
        
            e,v=np.linalg.eig(np.linalg.inv(F_hat+observed_information))
            objective=np.max(e)/np.min(e)
    
        return objective
        
    elif output=='covariance':
        
        return np.linalg.inv(F_hat+observed_information)          
             
                
                      
        
    
def expected_covariance(design_vector,length,theta,sigma,model,Obs_Fisher,measurable):
    
    C_Bz=design_vector[0]
    U=design_vector[1:]
    
    x0=np.array([C_Bz, 0.0])
    #V_theta=np.zeros(shape=(len(theta),len(theta)))
    
    F=expected_fisher(x0,U,length,theta,sigma,model,measurable)
    
    V_theta=np.linalg.inv(Obs_Fisher+F)
    
    return V_theta
    
def expected_covariance_novel(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model):
    
    U=design_vector[:-1]
    t=design_vector[-1]
    
    #V_theta=np.zeros(shape=(len(theta),len(theta)))
    
    F=expected_fisher(x0,U,t,theta,sigma,model,measurable)
    
    D=deviation(x0,U,t,theta,sigma,model,measurable,true_parameters,true_model)
    
    V_theta=np.linalg.inv(Obs_Fisher+F+D)
    
    return V_theta
    

    
def correlation(covariance):
    correlation=np.ndarray(shape=(len(covariance),len(covariance)))
    
    for i in range(0,len(covariance)):
        for j in range(0,len(covariance)):
            correlation[i,j]=covariance[i,j]/(np.sqrt(covariance[i,i]*covariance[j,j]))
    
    return correlation

        
        
def preliminary_guess(ranges,dataset,sigma,model,measurable,num_guess):
    
    ranges=np.transpose(ranges)
    
    objective_function=np.zeros(num_guess)
    theta=np.zeros(shape=(num_guess,len(ranges[0])))
    
    for i in range(0,num_guess):
        theta[i]=ranges[0]+(ranges[1]-ranges[0])*np.random.random(size=len(ranges[0]))
        objective_function[i]=loglikelihood(theta[i], dataset, sigma, model, measurable)
        
    return theta[objective_function.tolist().index(min(objective_function))]
    
def model_output(theta,x0,t,U,model,variable):
    
    X=integrate.odeint(model,x0,[0.0,t],args=(U,theta,))[1]
    
    return X[variable]
        
def deviation(x0,U,t,theta,sigma,model,measurable,true_parameters,true_model):
    
    deviation=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        deviation=deviation+((1/(sigma[i]**2))*(model_output(theta,x0,t,U,model,i)-model_output(true_parameters,x0,t,U,true_model,i))*nd.Hessian(model_output)(theta,x0,t,U,model,i))
    
    return deviation
    
def observed_deviation(x0,U,t,theta,sigma,model,measurable,measurements):
    
    deviation=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        deviation=deviation+((1/(sigma[i]**2))*(model_output(theta,x0,t,U,model,i)-measurements[i])*nd.Hessian(model_output)(theta,x0,t,U,model,i))
    
    return deviation
    
def observed_fisher(dataset,sigma,theta,model):
    
    observed_information=np.zeros(shape=(len(dataset),2,len(theta),len(theta)))
    
    for i in range(0,len(dataset)):
        observed_information[i,0]=expected_fisher(dataset[i,0],dataset[i,1],dataset[i,2],theta,sigma,model,dataset[i,4])
        observed_information[i,1]=observed_deviation(dataset[i,0],dataset[i,1],dataset[i,2],theta,sigma,model,dataset[i,4],dataset[i,3])
    
    
    observed_info_sensitivity=np.zeros(shape=(len(theta),len(theta)))
    observed_info_deviation=np.zeros(shape=(len(theta),len(theta)))

    for i in range(0,len(observed_information)):
        observed_info_sensitivity=observed_info_sensitivity+observed_information[i,0]
        observed_info_deviation=observed_info_deviation+observed_information[i,1]
    
    
    observed_fisher_information=observed_info_sensitivity+observed_info_deviation
    
    '''
    try:
        e, v = np.linalg.eig(observed_fisher_information)
        if np.all(e>0):
        
            return observed_fisher_information
        
        else:
            logger.warning('Computed non-semipostive definite information matrix')
            logger.warning('A large sample approximation is returned\n')
            return observed_info_sensitivity
    
    except LinAlgError:
        
        logger.warning('Error encountered in the diagonalisation of information matrix')
        logger.warning('A large sample approximation is returned\n')
        return observed_info_sensitivity
    '''
   
    return observed_info_sensitivity 
    
    


def excelsave(XX,YY,function,filename):
    
    wb=Workbook()

    ws=wb.active
    
    if XX.shape!=function.shape:
        function=function.reshape(XX.shape)
         
    
    for i in range(len(function)):
        for j in range(len(function)):
            ws.cell(row=i+2, column=1, value=(XX[i,0]/1))
            ws.cell(row=1, column=j+2, value=(YY[0,j]/1))
            ws.cell(row=i+2, column=j+2, value=function[i,j])
        
    wb.save(filename)
    
    return
    
def confidence_ellipsoid(total_covariance,significance,MLE,parameters):
    
    array_length=200
    covariance=np.zeros(shape=(2,2))
    
    for i in [0,1]:
        for j in [0,1]:
            covariance[i,j]=total_covariance[parameters[i],parameters[j]]
    
    eigenvalues,v=np.linalg.eig(covariance)
    
    if all([x>0 for x in eigenvalues]):
        eigenvalues=np.sqrt(eigenvalues)
    
        z=st.norm.ppf(significance)    
    
    
    
        ax = plt.subplot(111, aspect='equal')

        ell = Ellipse(xy=(MLE[parameters[0]],MLE[parameters[1]]),
                  width=eigenvalues[0]*z*2, height=eigenvalues[1]*z*2,
                  angle=np.rad2deg(np.arccos(v[0,0])))
        ell.set_facecolor('blue')
        ax.add_artist(ell)
    
    
        #plt.show()
    
    
        X=np.linspace(-eigenvalues[0]*z,eigenvalues[0]*z,array_length)
    
        Y=np.zeros(shape=(2*len(X),2))
    
        for i in range(0,len(X)):
            Y[i,0]=X[i]
            Y[i+len(X),0]=X[-(i+1)]
            Y[i,1]=(eigenvalues[1]*z)*np.sqrt(1-((Y[i,0]**2)/((eigenvalues[0]*z)**2)))
            Y[i+len(X),1]=-(eigenvalues[1]*z)*np.sqrt(1-((Y[i+len(X),0]**2)/((eigenvalues[0]*z)**2)))
            Y[i]=np.dot(v,Y[i])+[MLE[parameters[0]],MLE[parameters[1]]]
            Y[i+len(X)]=np.dot(v,Y[i+len(X)])+[MLE[parameters[0]],MLE[parameters[1]]]
    
    else:
        Y=np.zeros(shape=(2*array_length,2))
    
    return Y
    
def export_confidence_ellipsoids(covariance,significance,MLE,filename):
    
    wb=Workbook()
    
    ws1=wb.active
    
    ws1.title=('from_observed_covariance')
    
    index=1
    
    for i in range(0,len(MLE)):
        for j in range(i+1,len(MLE)):
            ell=confidence_ellipsoid(covariance,significance,MLE,[i,j])
            ws1.cell(row=1, column=index, value='Par. '+str(i+1))
            ws1.cell(row=1, column=index+1, value='Par. '+str(j+1))
            ws1.cell(row=2, column=index, value=MLE[i])
            ws1.cell(row=2, column=index+1, value=MLE[j])
            ws1.cell(row=3, column=index, value='Ell. Par. '+str(i+1))
            ws1.cell(row=3, column=index+1, value='Ell. Par. '+str(j+1))
            for k in range(len(ell)):
                ws1.cell(row=k+4, column=index, value=ell[k,0])
                ws1.cell(row=k+4, column=index+1, value=ell[k,1])
            index=index+3
    

    wb.save(filename)
    
    return
    

    
    
    
def normalised_to_design_space(coordinates, design_space):
    
    design_coordinates=np.zeros(len(coordinates))
    
    for i in range(len(coordinates)):
        
        design_coordinates[i]=design_space[i,0]+(design_space[i,1]-design_space[i,0])*coordinates[i]
        
    return design_coordinates
    
    
#Function t-test performs a t-test to assess the statistical quality of the parameter estimates 
   
def t_test(parameters, covariance, significance, dof):
    
    variances=np.diag(covariance)
    
    t_ref=st.t.interval(1-(2*(1-significance)),dof)[1]
    
    t_values=np.array(np.abs(parameters)/(st.t.interval(significance,dof)[1]*np.sqrt(variances)))
    
    return t_values, t_ref, np.all(t_values>t_ref)
    
def excelsave(array,filename):  #array is a numpy rectangular array; filename is a string of the type 'name_of_your_file.xlsx'
    
    wb=Workbook()

    ws=wb.active
    
    for i in range(len(array)):
        for j in range(len(array[0])):
            ws.cell(row=i+1, column=j+1, value=array[i][j])
        
    wb.save(filename)
    
    return
    
    
    
def parameter_space_transformation(contraction_coefficient, covariance, parameters):
    
   
    #Compute the eigenvalues of the covariance
    
    eigenvalues = np.linalg.eig(covariance)[0]

    #Compute the matrix with the eigenvectors of the covariance
    
    first_rotation_matrix = np.linalg.eig(covariance)[1]
    
    #evaluates the average eigenvalue to assess whether a certain eigenvalue shall be reduced or increased
    
    average_eigenvalue = np.sum(np.sqrt(eigenvalues))/len(eigenvalues)
    
    #computes a reshape matrix for the given contraction coefficient (contraction coefficient is always assumed 1)
    
    reshape_matrix =np.diag((np.sqrt(eigenvalues)/average_eigenvalue)*contraction_coefficient)
    
    #Compute a partial transformation which aligns the ellipsoids axes to
    #the axes of the parameter space and scales all the ellipsoid axes to an average value
    #after transformation_1 the condition number is theoretically equal to 1
    
    transformation_1=np.dot(first_rotation_matrix, reshape_matrix)
    
    #Applies transformation_1 to compute the value of parameters after the partial transformation
      
    new_parameters=np.dot(np.linalg.inv(transformation_1), parameters)
    
    #Compute the distance of the parameters from the origin after the partial transformation
    #This transformation is required to change the scales of the parameters to something more treatable by the algorithms
    
    target_parameter_value=100.0 #<-assumed by default
    
    distance=np.sqrt(np.sum(new_parameters**2))/np.sqrt((target_parameter_value**2)*len(parameters))
        
    #Apply the distance to compute a second partial transformation after which the parameters lie on the
    #hypersphere with radius sqrt(N_theta*target_parameter_value**2)
    
    transformation_2=distance*transformation_1
    
    #A last partial transformation is computed to rotate the parameter space
    #so that all the parameters have magnitude 100.0    

    target_angle=math.atan2(1.0, 1.0)
    
    
    second_rotation_matrix=np.eye(len(parameters))
    
    #The following code computes the rotation matrix R by applying elementary rotations
    #It has been validated only in the 2-parameters case
    
    for i in range(len(parameters)):
        
        for j in range(i+1,len(parameters)):
            
             elementary_rotation=np.zeros((len(parameters), len(parameters)))
             
              
             angle=math.atan2(new_parameters[j], new_parameters[i])
             
              
             rotation=angle-target_angle
             
             elementary_rotation[i,i]=np.cos(rotation)
             elementary_rotation[j,j]=np.cos(rotation)
             elementary_rotation[i,j]=-np.sin(rotation)
             elementary_rotation[j,i]=np.sin(rotation)
             
             second_rotation_matrix=np.dot(elementary_rotation, second_rotation_matrix)
              
    #Compute the total transformation matrix
    
    transformation=np.dot(transformation_2, second_rotation_matrix)
    
    #Apply the total transformation to the initial value of parameters
    #to compute the parameter values in the transformed parameter space
    #Theoretically, the array new_parameters_2 should contain only values equal to target_parameter_value
   
    new_parameters_2=np.dot(np.linalg.inv(transformation), parameters)
    
    return new_parameters_2, transformation
    
    
def evaluate_statistics_in_transformed_space(robust_parameters, covariance_robust_parameters, transformation, dataset, measurable, output, filename=None, true_parameters=None):
    '''
    np.random.seed(0)
    samples=np.random.multivariate_normal(robust_parameters, covariance_robust_parameters, 500)
    
    new_samples=np.array([np.dot(transformation, samples[i]) for i in range(len(samples))])
    
    covariance_original_parameters=np.cov(np.transpose(new_samples))
    '''
    
    original_parameters=np.dot(transformation, robust_parameters)
    covariance_original_parameters=np.matmul(transformation, np.matmul(covariance_robust_parameters, np.transpose(transformation)))

    eigenvalues, eigenvectors = np.linalg.eig(covariance_original_parameters)
        
    condition_number=np.sqrt(np.max(eigenvalues)/np.min(eigenvalues))



    
    if output=='stats':
        
        logger.info('Parameters are: {}'.format(original_parameters))
        logger.info('95% confidence interval: +- {}'.format(2.0*np.sqrt(np.diag(covariance_original_parameters))))
        logger.info('Computed covariance: {}'.format(np.array_repr(covariance_original_parameters).replace('\n', '')))
        logger.info('The correlation of parameters is: {}'.format(correlation(covariance_original_parameters)[0,1]))
        logger.info('t-statistics of parameters are: {}'.format(t_test(original_parameters, covariance_original_parameters, 0.95, len(dataset)*len(measurable)-len(robust_parameters))))
        logger.info('Condition number: {}\n'.format(condition_number))
        
        if true_parameters!=None:
            distance=true_parameters-original_parameters
            statistics=np.matmul(distance, np.matmul(np.linalg.inv(covariance_original_parameters), np.transpose(distance)))
            logger.info('The p-value of the true parameters given the computed distribution is: {}%\n'.format((1-st.chi2.cdf(statistics,len(robust_parameters)))*100))
        
        return original_parameters, t_test(original_parameters, covariance_original_parameters, 0.95, len(dataset)*len(measurable)-len(robust_parameters)), correlation(covariance_original_parameters)[0,1]
     
        #Compute the p-value of the dataset, given the distribution of the parameters
        
        #dataset_sample=simulate_dataset(dataset, model, new_samples)
        
        
        
 
                
                
        
    elif output=='ellipsoid':
        
        export_confidence_ellipsoids(covariance_original_parameters, 0.95, original_parameters, filename)
        
    return
    
