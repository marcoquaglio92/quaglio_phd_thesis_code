import math
import numpy as np
from scipy import integrate
from scipy.optimize import minimize
import scipy.stats as st
from matplotlib.patches import Ellipse
import matplotlib.pyplot as plt
import numdifftools as nd
from openpyxl import Workbook, load_workbook
import copy
import time as time





def differential_model(x,t,u,theta, equations):
    """
    General model in the form of a system of ODEs. Structured to be called by the scipy odeint package.
    
    Inputs:
        x: array of state values
        t: time value
        u: array of time-invariant inputs
        theta: array of model parameters,
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'

    Output: Array of dim(x) with the numerical values for the first order derivatives at the conditions specified by the inputs 
    
    """
           
    dxdt=eval(equations)
    
    return dxdt


def measurement(x0,U,t,sigma,theta,equations,measurable):
    '''
    Generates a sample in-silico by integrating the system of ODEs defined by the "equations" input
    
    Inputs:
        x0: array of initial states
        U: array of time-invariant system inputs
        t: scalar sampling time
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        theta: array of parameters used to integrate the assumed system model
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        measurable: array of integers that defines which states are measured, e.g. if x=[x_0,x_1,x_2], but only x_0 and x_2 are measured, then measurable=np.array([0,2])
        
    Output:
        dataset type element storing all the information about the sample -> [initial states, inputs, sampling time, measurements, measured states]   
    '''
    
    X=integrate.odeint(differential_model,x0,[0.0,t],(U, theta, equations))
    measurement=np.abs(X[1]+np.array(sigma)*np.random.normal(0.0,np.ones(len(sigma))))
    return [np.array(x0),np.array(U),t,np.array(measurement),measurable]


#Function loglikelihood fits a model to the experimental data. It was preferred to integrate the model for every measurement singularly sacrifying computational efficiency

def loglikelihood(theta, dataset, sigma, equations):
    '''
    Computes the residual term of the loglikelihood function for an ODEs system
    
    Inputs:
        theta: array of model parameters
        dataset: dataset type array 
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
                    
    Output:
        1/2*sum of normalised squared residuals
    '''
    
    
    residuals=np.zeros(shape=(len(dataset),len(sigma)))
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=[dataset[i][3] for i in range(0,len(dataset))]
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(differential_model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], theta, equations))
        predictions[i]=X[1]
    
    residuals=((measurements-predictions)/sigma)**2
    
    objective=0
    
    for i in range(0,len(dataset)):
        objective=objective+np.sum(residuals[i,dataset[i][4]])
            
    return 0.5*objective
      

def gradient_loglikelihood(theta, dataset, sigma, equations, epsilon=0.000001):
    '''
    Computes the gradient in the parameter space of the loglikelihood function for an ODEs system
    
    Inputs:
        theta: array of model parameters
        dataset: dataset type array 
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        epsilon: scalar value defining the perturbation of parameters for the evaluation of the sensitivity
        
    Output:
        gradient - array of partial derivatives
    '''
    
    
    unperturbed_loglikelihood=loglikelihood(theta, dataset, sigma, equations)*np.ones(len(theta))
    
    
    
    perturbed_theta=np.repeat([theta],len(theta),axis=0)+epsilon*np.eye(len(theta))*np.maximum.reduce([theta,np.ones(len(theta))*0.1])
    
    #print(epsilon*np.maximum.reduce([theta,np.ones(len(theta))*0.1]))
    
    perturbed_loglikelihood=np.zeros(len(theta))
    
    for i in range(0, len(theta)):
        
        perturbed_loglikelihood[i]=loglikelihood(perturbed_theta[i], dataset, sigma, equations)
        
    
    #print(perturbed_loglikelihood-unperturbed_loglikelihood)
    
    gradient=np.divide(perturbed_loglikelihood-unperturbed_loglikelihood, epsilon*np.maximum.reduce([theta,np.ones(len(theta))*0.1]))

    return gradient


 
  
def experiment(x0,U,st,sigma,theta,equations,measurable):
    '''
    Performs an in-silico experiment given initial states, time-invariant control input variables, and an array of sampling times 
    
    Inputs:
        x0: array of initial states
        U: array of time-invariant system inputs
        st: array of sampling times
        sigma: array of standard deviations for the uncorrelated measurement noise associated to the state variables
        theta: array of parameters used to integrate the assumed system model
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        measurable: array of integers that defines which states are measured, e.g. if x=[x_0,x_1,x_2], but only x_0 and x_2 are measured, then measurable=np.array([0,2])
        
    Output:
        dataset type element storing all the information about the samples -> [initial states, inputs, sampling time, measurements, measured states]   
  
    '''
    dataset=np.array([measurement(x0,U,st[0],sigma,theta,equations,measurable)])
    for i in range(1,len(st)):
        dataset=np.append(dataset,[measurement(x0,U,st[i],sigma,theta,equations,measurable)],axis=0)
        
    return dataset

#sensitivity evaluates the gradient of the measurable model output in the parameter space
def sensitivity(x0,U,t,theta,equations,epsilon):
    '''
    Computes the N_{theta}xN_{states} sensitivity matrix related to a single sample 
    
    Inputs:
        x0: array of initial states
        U: array of time-invariant system inputs
        t: scalar sampling time
        theta: array of parameters used to integrate the assumed system model
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        epsilon: scalar value defining the perturbation of parameters for the evaluation of the sensitivities
        
    Output:
        N_{theta}xN_{states} sensitivity matrix 
  
    '''
    #epsilon=0.0001
    sensitivity_matrix=np.ndarray(shape=(len(theta),len(x0)))
    perturbed_X=np.ndarray(shape=(len(theta),len(x0)))
    unperturbed_X=integrate.odeint(differential_model,x0,[0.0,t],args=(U,theta,equations))[1]
    unperturbed_X=np.repeat([unperturbed_X],len(theta),axis=0)
    perturbed_theta=np.repeat([theta],len(theta),axis=0)+epsilon*np.eye(len(theta))*np.maximum.reduce([theta,np.ones(len(theta))*0.1])
    for i in range(0,len(theta)):
        perturbed_X[i]=integrate.odeint(differential_model,x0,[0.0,t],args=(U,perturbed_theta[i],equations))[1]
    
    
    
    sensitivity_matrix=np.transpose(np.divide(np.transpose(np.subtract(perturbed_X,unperturbed_X)), epsilon*np.maximum.reduce([theta,np.ones(len(theta))*0.1])))
    
    return sensitivity_matrix
            


def expected_fisher(x0,U,t,theta,sigma,equations,measurable, epsilon):
    '''
    calculates the N_{theta}xN_{theta} Fisher information matrix related to a single sample
    
    Inputs:
        x0: array of initial states
        U: array of time-invariant system inputs
        t: scalar sampling time
        theta: array of parameters used to integrate the assumed system model
        sigma: array of standard deviations for the uncorrelated measurement noise associated to the state variables
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        measurable: array of integers that defines which states are measured, e.g. if x=[x_0,x_1,x_2], but only x_0 and x_2 are measured, then measurable=np.array([0,2])
        epsilon: scalar value defining the perturbation of parameters for the evaluation of the sensitivities
        
    Output:
        N_{theta}xN_{theta} information matrix related to a single sample 
    
    '''
    Q=sensitivity(x0,U,t,theta,equations,epsilon)
    Fisher=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        Fisher=Fisher+(1/(sigma[i]**2))*np.outer(Q[:,i],Q[:,i])
        
    return Fisher
    
    

 
    
def correlation(covariance):
    correlation=np.ndarray(shape=(len(covariance),len(covariance)))
    
    for i in range(0,len(covariance)):
        for j in range(0,len(covariance)):
            correlation[i,j]=covariance[i,j]/(np.sqrt(covariance[i,i]*covariance[j,j]))
    
    return correlation
    
    

    
def observed_fisher(dataset, sigma, theta, equations, epsilon=0.000001):
    '''
    Computes the large sample approximation of the observed Fisher information matrix 
    
    Inputs:
        
        dataset: dataset type numpy array
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        theta: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        epsilon: quantifies the relative perturbation used to calculate gradients and sensitivities
    '''
    observed_information=np.zeros(shape=(len(dataset),2,len(theta),len(theta)))
    
    for i in range(0,len(dataset)):
        observed_information[i,0]=expected_fisher(dataset[i,0],dataset[i,1],dataset[i,2],theta,sigma,equations,dataset[i,4], epsilon)
        #observed_information[i,1]=observed_deviation(dataset[i,0],dataset[i,1],dataset[i,2],theta,sigma,model,dataset[i,4],dataset[i,3])
    
    
    observed_info_sensitivity=np.zeros(shape=(len(theta),len(theta)))
    #observed_info_deviation=np.zeros(shape=(len(theta),len(theta)))

    for i in range(0,len(observed_information)):
        observed_info_sensitivity=observed_info_sensitivity+observed_information[i,0]
        #observed_info_deviation=observed_info_deviation+observed_information[i,1]
    
    
    #observed_fisher_information=observed_info_sensitivity+observed_info_deviation
    

   
    return observed_info_sensitivity 
    
    

    
def confidence_ellipsoid(total_covariance,significance,MLE,parameters):
    
    array_length=200
    covariance=np.zeros(shape=(2,2))
    
    for i in [0,1]:
        for j in [0,1]:
            covariance[i,j]=total_covariance[parameters[i],parameters[j]]
    
    eigenvalues,v=np.linalg.eig(covariance)
    
    if all([x>0 for x in eigenvalues]):
        eigenvalues=np.sqrt(eigenvalues)
    
        z=st.norm.ppf(significance)    
    
    
    
        #ax = plt.subplot(111, aspect='equal')

        #ell = Ellipse(xy=(MLE[parameters[0]],MLE[parameters[1]]),
        #          width=eigenvalues[0]*z*2, height=eigenvalues[1]*z*2,
        #          angle=np.rad2deg(np.arccos(v[0,0])))
        #ell.set_facecolor('blue')
        #ax.add_artist(ell)
    
    
        #plt.show()
    
    
        X=np.linspace(-eigenvalues[0]*z,eigenvalues[0]*z,array_length)
    
        Y=np.zeros(shape=(2*len(X),2))
    
        for i in range(0,len(X)):
            Y[i,0]=X[i]
            Y[i+len(X),0]=X[-(i+1)]
            Y[i,1]=(eigenvalues[1]*z)*np.sqrt(1-((Y[i,0]**2)/((eigenvalues[0]*z)**2)))
            Y[i+len(X),1]=-(eigenvalues[1]*z)*np.sqrt(1-((Y[i+len(X),0]**2)/((eigenvalues[0]*z)**2)))
            Y[i]=np.dot(v,Y[i])+[MLE[parameters[0]],MLE[parameters[1]]]
            Y[i+len(X)]=np.dot(v,Y[i+len(X)])+[MLE[parameters[0]],MLE[parameters[1]]]
    
    else:
        Y=np.zeros(shape=(2*array_length,2))
    
    return Y
    
def export_confidence_ellipsoids(covariance,significance,MLE,filename):
    '''
    The function exports confidence ellipsoids with a user defined level of significance
    
    Inputs:
        covariance: covariance matrix of the parameter estimates
        significance: significance level of the ellipsoid, e.g. 0.95
        MLE: array-type object representing the maximum likelihood estimate
        filename: string-type object, e.g. 'confidence_ellipsoid.xlsx'
        
    Output:
        generates an excel file with the sequence of points required to plot the ellipsoid.
        the file is generated in the working directory  
         
    
    '''
    wb=Workbook()
    
    ws1=wb.active
    
    ws1.title=('from_observed_covariance')
    
    index=1
    
    for i in range(0,len(MLE)):
        for j in range(i+1,len(MLE)):
            ell=confidence_ellipsoid(covariance,significance,MLE,[i,j])
            ws1.cell(row=1, column=index, value='Par. '+str(i+1))
            ws1.cell(row=1, column=index+1, value='Par. '+str(j+1))
            ws1.cell(row=2, column=index, value=MLE[i])
            ws1.cell(row=2, column=index+1, value=MLE[j])
            ws1.cell(row=3, column=index, value='Ell. Par. '+str(i+1))
            ws1.cell(row=3, column=index+1, value='Ell. Par. '+str(j+1))
            for k in range(len(ell)):
                ws1.cell(row=k+4, column=index, value=ell[k,0])
                ws1.cell(row=k+4, column=index+1, value=ell[k,1])
            index=index+3
    

    wb.save(filename)
    
    return
    

    
    
#Function t-test performs a t-test to assess the statistical quality of the parameter estimates 
   
def t_test(parameters, covariance, significance, dof):
    '''
    Perform a t-test for parameter significance.
    
    Input:
        parameters: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
        covariance: numpy array with the covariance of the parameter estimates,
        significance: significance level of the test
        dof: degree of freedom of t-distribution
        
    Output:
        
        array of t_values computed according to gPROMS definition, t_ref, array of test oucomes True/False for each parameter
    
    '''
    variances=np.diag(covariance)
    
    t_ref=st.t.interval(1-(2*(1-significance)),dof)[1]
    
    t_values=np.array(np.abs(parameters)/(st.t.interval(significance,dof)[1]*np.sqrt(variances)))
    
    return t_values, t_ref, np.all(t_values>t_ref)
    

    

    



def loglikelihood_higher_dimensional_space(theta, parameter_to_diagnose, dataset, sigma, equations):
    '''
    Computes the residual term of the loglikelihood function under a parametrisation that assumes that parameter_to_diagnose is a function of the experimental conditions
    The function is called to compute the Lagrange multipliers statistic required to compute the model modification index
    
    Input:
        theta: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
              
        parameter_to_diagnose: an integer 'i' referring to theta[i], the parameter under diagnosis
              
        dataset: dataset type numpy array
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
                    
    Output: (1/2)*sum of squared residuals under a new parametrisation for model diagnosis 
            
    '''
    
    
    distributed_parameter=theta[0:len(dataset)]
    
    constant_parameters=theta[len(dataset):]
    
    #construction of parameter sets
    
    parameter_sets=np.insert(np.repeat(np.array([constant_parameters]), len(dataset), axis=0), parameter_to_diagnose, distributed_parameter, axis=1)
 
    
    residuals=np.zeros(shape=(len(dataset),len(sigma)))
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=[dataset[i][3] for i in range(0,len(dataset))]
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(differential_model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], parameter_sets[i],equations))
        predictions[i]=X[1]
    
    residuals=((measurements-predictions)/sigma)**2
    
    objective=0
    
    for i in range(0,len(dataset)):
        objective=objective+np.sum(residuals[i,dataset[i][4]])
            
    return 0.5*objective



def compute_fisher_in_high_dimensional_space(theta, parameter_to_diagnose, dataset, sigma, equations, epsilon=0.0001):
    '''
    Computes the "N_samples+N_theta-1 x N_samples+N_theta-1" fisher information matrix under a parametrisation that assumes that parameter_to_diagnose is a function of the experimental conditions
    The function is called to compute the Lagrange multipliers statistic required to compute the model modification index
    
    Input:
        theta: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
              
        parameter_to_diagnose: an integer 'i' referring to theta[i], the parameter under diagnosis
              
        dataset: dataset type numpy array
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
                    
        epsilon: quantifies the relative perturbation used to calculate gradients and sensitivities   
        
    Output: "N_samples+N_theta-1 x N_samples+N_theta-1" fisher information matrix
            
    '''
    fisher=np.zeros((len(dataset)+len(theta)-1,len(dataset)+len(theta)-1))
    
    for i in range(0, len(dataset)):
        
        sensitivity_in_restricted_space=sensitivity(dataset[i,0], dataset[i,1], dataset[i,2], theta, equations, epsilon)[:, dataset[i,4]]

        sensitivity_in_higher_dimensional_space=np.zeros((len(dataset)+len(theta)-1, len(dataset[0,4])))
        
        sensitivity_in_higher_dimensional_space[i]=sensitivity_in_restricted_space[parameter_to_diagnose]
        
        sensitivity_in_higher_dimensional_space[-len(theta)+1:]=np.delete(sensitivity_in_restricted_space, parameter_to_diagnose, axis=0)
    
        fisher=fisher+np.matmul(sensitivity_in_higher_dimensional_space, np.matmul(np.linalg.inv(np.diag(sigma[dataset[i][4]])**2), np.transpose(sensitivity_in_higher_dimensional_space)))        
        
        
    return fisher
   

def gradient_likelihood_in_high_dimensional_space(parameter_estimates, parameter_to_diagnose, dataset, sigma, equations, method='default', epsilon=0.0001):
    '''
    Computes the gradient of the loglokielihood function under a parametrisation that assumes that parameter_to_diagnose is a function of the experimental conditions
    The function is called to compute the Lagrange multipliers statistic required to compute the model modification index
    
    Input:
        parameter_estimates: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
              
        parameter_to_diagnose: an integer 'i' referring to theta[i], the parameter under diagnosis
              
        dataset: dataset type numpy array
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
                    
        method:
            'default' (default): calculates gradients applying a perturbation proportional to epsilon 
        epsilon: quantifies the relative perturbation used to calculate gradients and sensitivities   
        
    Output: log-likelihood gradient under the new parametrisation
            
    '''
    
    extended_parameter_set=np.append(np.ones((len(dataset)))*parameter_estimates[parameter_to_diagnose], np.delete(parameter_estimates, parameter_to_diagnose))

    if method=='numdifftool':
        
        gradient=nd.Gradient(loglikelihood_higher_dimensional_space)(extended_parameter_set, parameter_to_diagnose, dataset, sigma, equations)
        
        return gradient
    
    
    gradient=np.zeros(len(dataset)+len(parameter_estimates)-1)

    
    perturbation=epsilon*extended_parameter_set
    
    
    unperturbed_likelihood_high_dimensional_space=loglikelihood_higher_dimensional_space(extended_parameter_set, parameter_to_diagnose, dataset, sigma, equations)
    
    
    perturbed_parameter_sets=np.repeat([extended_parameter_set],len(extended_parameter_set),axis=0)+epsilon*np.eye(len(extended_parameter_set))*extended_parameter_set
    
    #print(perturbation)
    
    for i in range(0, len(gradient)):
           
        gradient[i]=(loglikelihood_higher_dimensional_space(perturbed_parameter_sets[i], parameter_to_diagnose, dataset, sigma, equations)-unperturbed_likelihood_high_dimensional_space)/perturbation[i]
    
    #print(gradient)
    return gradient
    


def lagrange_multiplier_diagnosis(equations, parameter_estimates, dataset, sigma, method='univariate', gradient='default', parameter_to_diagnose='all', epsilon=0.0001):
    '''
    Performs a Lagrange multipliers test to disprove the hypothesis that a given parameter is a state-independent constant
    
    Input:
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
        
        parameter_estimates: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
        dataset: dataset type numpy array
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        method:
            'univariate' (default) treats the parameters that are not under diagnosis as fixed coefficients
            'multivariate' considers correlation among all model parameters
            
        gradient:
            'default' (default): calculates gradients applying a perturbation proportional to epsilon 
            'numdifftools': computes the gradient of the log-likelihood function using the package numdifftools; 
            
        parameter_to_diagnose: 
            'all' (default): evaluates the Lagrange multipliers statistic for all model parameters
            otherwise, if an integer i is specified, the Lagrange multipliers statistic is computed only on parameter theta[i]
        
        epsilon: quantifies the relative perturbation used to calculate gradients and sensitivities   
        
    Output: Lagrange multipler statistic, 95% reference value
            
    '''
    
    if parameter_to_diagnose=='all':
        
        score_statistics=np.zeros(len(parameter_estimates))
        
        for parameter_under_diagnosis in range(0, len(parameter_estimates)):
            
            score=gradient_likelihood_in_high_dimensional_space(parameter_estimates, parameter_under_diagnosis, dataset, sigma, equations, gradient, epsilon)
            
            expected_fisher=compute_fisher_in_high_dimensional_space(parameter_estimates, parameter_under_diagnosis, dataset, sigma, equations, epsilon)
            
            if method=='univariate':
                score_statistics[parameter_under_diagnosis]=np.matmul(score[:len(dataset)], np.matmul(np.linalg.inv(expected_fisher[:len(dataset),:len(dataset)]), np.transpose(score[:len(dataset)])))
            elif method=='multivariate':
                score_statistics[parameter_under_diagnosis]=np.matmul(score, np.matmul(np.linalg.inv(expected_fisher), np.transpose(score)))
                
            else:
                print('invalid method selected: only unvariate and multivariate are allowed')
                
                return
                
        return score_statistics, st.chi2.ppf(0.95, len(dataset)-1)
            
    else: 
            
        score=gradient_likelihood_in_high_dimensional_space(parameter_estimates, parameter_to_diagnose, dataset, sigma, equations, gradient)

        expected_fisher=compute_fisher_in_high_dimensional_space(parameter_estimates, parameter_to_diagnose, dataset, sigma, equations)
            
        if method=='univariate':
            score=np.matmul(score[:len(dataset)], np.matmul(np.linalg.inv(expected_fisher[:len(dataset),:len(dataset)]), np.transpose(score[:len(dataset)])))
        elif method=='multivariate':
            score=np.matmul(score, np.matmul(np.linalg.inv(expected_fisher), np.transpose(score)))
                
        else:
            print('invalid method selected: only unvariate and multivariate are allowed')
                
            return
                
        return score, st.chi2.ppf(0.95, len(dataset)-1)
        

   
      
#Evaluation of first order effects with Lagrange multipliers test

def first_order_effects_on_parameters(equations, parameter_estimates, dataset, sigma, parameters_to_diagnose=[], considered_effects=[], approach='multivariate', method=None, output=None, epsilon=0.00000000001):
    '''
    The function automatically generates model structures by evolving parameters to diagnose into first order response surfaces theta_i -> theta_i + (theta_j * effect)
    and computes a Lagrange multipliers statistic with the aim of challenging the hypothesis theta_j=0 against the hypothesis that theta_j differs from 0
    
    Inputs: 
        equations: string-like object containing the expression of the first order derivatives
                    e.g. 'np.array([    ((theta[0]*x[1]/(theta[1]*x[0]+x[1]))-u[0]-theta[3])*x[0],        -((theta[0]*x[1]/(theta[1]*x[0]+x[1]))*x[0])/theta[2]+u[0]*(u[1]-x[1])    ])'
                    
        parameter_estimates: array of parameters which represents the current maximum likelihood estimate for the model under diagnosis
        dataset: dataset type numpy array
        sigma: array of standard deviations for the uncorrelated measurement noise of the measured quantities involved in a sample
        parameters_to_diagnose: list/array of integers representing the model parameters the user is willing to diagnose
        considered_effects: array of strings storing the effects the user is willing to consider, e.g. np.array(['x[0]', 'x[1]', '1/x[0]'])
        approach: 
            'multivariate' (considers the full information matrix extended with theta_j); 
            'univariate' (considers only parameter theta_j associated to the considered effect in the computation of the Lagrange multipliers statistic)
            'zero-first order' (considered only parameters theta_i and theta_j involved in the evolved model bit)
            
        method: 
            'numdifftools' computes the gradient of the log-likelihood function using the package numdifftools; 
             None (default) computes the gradient with a perturbation proportional to epsilon
        output: 
            'print' prints the generated model structure and the computed effect relevance index ERI= Lagrange statistic / chi2(95%)
             None (default) the function does not print anything at console
        epsilon: quantifies the relative perturbation used to calculate gradients and sensitivities
        
    Output: first_order_effects (dictionary) which contains an ERI computed for each considered effect and for each diagnosed parameter
                 
    
    '''
    no_states=len(dataset[0,0])
    no_inputs=len(dataset[0,1])
    
    #If no effect is specified, the analysis is performed for all states and inputs and their inverse
    
    if len(considered_effects)==0:
        
        considered_effects=[]
        
        for i in range(0,no_states):
            
            considered_effects.append('x['+str(i)+']')
            considered_effects.append('1/x['+str(i)+']')
            
        for i in range(0,no_inputs):
            
            considered_effects.append('u['+str(i)+']')
            considered_effects.append('1/u['+str(i)+']')
    
    #If no parameter is specified, the analysis is performed for all parameters
    
    if len(parameters_to_diagnose)==0:
        
        parameters_to_diagnose=np.array(range(len(parameter_estimates)))
        
        #for i in range(0,len(parameter_estimates)):
            
            #parameters_to_diagnose.append('theta['+str(i)+']')
            
            
    
    #A dictionary is inisialised to store the first order effects
    
    first_order_effects={}
    
    extended_parameter_set=np.append(parameter_estimates, 0.0)
    
    
    for i in parameters_to_diagnose:
        
        first_order_effects.update({'theta['+str(i)+']':{}})       
               
        for effect in considered_effects:
        
            new_equations=equations.replace('theta['+str(i)+']', '(theta['+str(i)+']+theta['+str(len(parameter_estimates))+']*'+'('+effect+'))')
            #new_equations=equations.replace(parameter, '('+parameter+'*('+effect+')**theta['+str(len(parameter_estimates))+'])')
            #new_equations=equations.replace(parameter, '(theta['+str(len(parameter_estimates))+']*'+'('+effect+'))')
            
            
            
            #Evaluate the gradient of the log-likelihood function with the new equations
            
            
            if method=='numdifftools':
                score=nd.Gradient(loglikelihood)(extended_parameter_set, dataset, sigma, new_equations, epsilon)
            else:
                score=gradient_loglikelihood(extended_parameter_set, dataset, sigma, new_equations, epsilon)
            
            #Evaluate expected Fisher information at constrained estimate
            
            information_matrix=observed_fisher(dataset, sigma, extended_parameter_set, new_equations, epsilon)
            
            if approach=='univariate':
                score_statistics=score[-1]*((information_matrix[-1,-1])**(-1))*score[-1]/st.chi2.ppf(0.95, 1)
            elif approach=='multivariate':
                try:
                    score_statistics=np.matmul(score, np.matmul(np.linalg.inv(information_matrix), np.transpose(score)))/st.chi2.ppf(0.95, 1)
                except:
                    print('Singular information matrix encountered')
                    score_statistics='Singular information matrix'
                    
            elif approach=='zero-first order':
                #print(score)
                score_statistics=np.matmul(score[np.array([i,-1])], np.matmul(np.linalg.inv(information_matrix[np.ix_([i,-1],[i,-1])]), np.transpose(score[np.array([i,-1])])))/st.chi2.ppf(0.95, 1)
                #print(score[np.array([i,-1])])
            else:
                print('invalid approach selected: only unvariate and multivariate are allowed')
            
            if output=='print':
                #print(new_equations)
                print('Effect '+ effect + ' on Parameter theta['+str(i)+']:{0:.2f}'.format(score_statistics))
            
            first_order_effects['theta['+str(i)+']'].update({effect:score_statistics})
        
        

    
    
    return first_order_effects
    