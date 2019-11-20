import math
import numpy as np
from scipy import integrate
from scipy.optimize import minimize
import scipy.stats as st
import matplotlib.pyplot as plt
import numdifftools as nd
from openpyxl import Workbook

#Incorrect model structure

def powerlaw_ethanol_fixed_energy(n,L,U,theta):
    
    A=np.array([16.5,-4.79])
    B=np.array([-9134.6,4386.0])
    
    #(A1, Ea, A2, Ea2, A3, Ea3)=theta
    
    Ea1, Ea2 = (36.25,12.95)
    
    (K220_1, K220_2)=theta
    
    stoichiometry=np.array([[-1,+1,0,+1],[-1,-1,+1,+1]])
    
    (Ptot,T)=U    #residence time and temperature
     
    Tref=273.15+220.0
    
    (Et, Ac, EA, H2)=n
    
    #Arrhenius
    #K1=A1*np.exp(-Ea1/R*T)
    #K2=A2*np.exp(-Ea2/R*T)
    #K3=A3*np.exp(-Ea3/R*T)
    
    Tref=220.0+273.15
       
    K1=K220_1*np.exp((Ea1/R)*((1/Tref)-(1/T)))
    K2=K220_2*np.exp((Ea2/R)*((1/Tref)-(1/T)))
    #K3=K220_3*np.exp((Ea3/R)*((1/Tref)-(1/T)))   
             
    Keq1=np.exp(A[0]+B[0]/T)
    
    Keq2=np.exp(A[1]+B[1]/T)
    
    P_Et=Et*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_Ac=Ac*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_EA=EA*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_H2=H2*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    rate=np.array([  K1*P_Et*(1-(1/Keq1)*(P_H2*P_Ac/P_Et)),
                     K2*P_Et*P_Ac*(1-(1/Keq2)*(P_EA*P_H2/(P_Et*P_Ac)))])
    
    #differential equations
    
    
    
    dndt=np.array(W*np.dot(np.transpose(stoichiometry),rate))
    return dndt



def powerlaw_ethanol(n,L,U,theta):
    
    A=np.array([16.5,-4.79])
    B=np.array([-9134.6,4386.0])
    
    #(A1, Ea, A2, Ea2, A3, Ea3)=theta
    
    (K220_1, Ea1, K220_2, Ea2)=theta
    
    stoichiometry=np.array([[-1,+1,0,+1],[-1,-1,+1,+1]])
    
    (Ptot,T)=U    #residence time and temperature
     
    Tref=273.15+220.0
    
    (Et, Ac, EA, H2)=n
    
    #Arrhenius
    #K1=A1*np.exp(-Ea1/R*T)
    #K2=A2*np.exp(-Ea2/R*T)
    #K3=A3*np.exp(-Ea3/R*T)
    
    Tref=220.0+273.15
       
    K1=K220_1*np.exp((Ea1/R)*((1/Tref)-(1/T)))
    K2=K220_2*np.exp((Ea2/R)*((1/Tref)-(1/T)))
    #K3=K220_3*np.exp((Ea3/R)*((1/Tref)-(1/T)))   
             
    Keq1=np.exp(A[0]+B[0]/T)
    
    Keq2=np.exp(A[1]+B[1]/T)
    
    P_Et=Et*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_Ac=Ac*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_EA=EA*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_H2=H2*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    rate=np.array([  K1*P_Et*(1-(1/Keq1)*(P_H2*P_Ac/P_Et)),
                     K2*P_Et*P_Ac*(1-(1/Keq2)*(P_EA*P_H2/(P_Et*P_Ac)))])
    
    #differential equations
    
    
    
    dndt=np.array(W*np.dot(np.transpose(stoichiometry),rate))
    return dndt



def powerlaw_ethanol_three(n,L,U,theta):
    
    A=np.array([16.5,-4.79])
    B=np.array([-9134.6,4386.0])
    
    #(A1, Ea, A2, Ea2, A3, Ea3)=theta
    
    (K220_1, Ea1, K220_2, Ea2, K220_3, Ea3)=theta
    
    stoichiometry=np.array([[-1,+1,0,+1],[-1,-1,+1,+1],[0,-2,0,0]])
    
    (Ptot,T)=U    #residence time and temperature
     
    Tref=273.15+220.0
    
    (Et, Ac, EA, H2)=n
    
    #Arrhenius
    #K1=A1*np.exp(-Ea1/R*T)
    #K2=A2*np.exp(-Ea2/R*T)
    #K3=A3*np.exp(-Ea3/R*T)
    
    Tref=220.0+273.15
       
    K1=K220_1*np.exp((Ea1/R)*((1/Tref)-(1/T)))
    K2=K220_2*np.exp((Ea2/R)*((1/Tref)-(1/T)))
    K3=K220_3*np.exp((Ea3/R)*((1/Tref)-(1/T)))   
             
    Keq1=np.exp(A[0]+B[0]/T)
    
    Keq2=np.exp(A[1]+B[1]/T)
    
    P_Et=Et*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_Ac=Ac*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_EA=EA*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_H2=H2*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    rate=np.array([  K1*P_Et*(1-(1/Keq1)*(P_H2*P_Ac/P_Et)),
                     K2*P_Et*P_Ac*(1-(1/Keq2)*(P_EA*P_H2/(P_Et*P_Ac))),
                     K3*P_Ac**2])
    
    #differential equations
    
    
    
    dndt=np.array(W*np.dot(np.transpose(stoichiometry),rate))
    return dndt


def monod(x,t,U,theta):
    
    (u1,u2)=U
    #differential equations
    dxdt=np.zeros((len(x),1))
    
    r=theta[0]*x[1]/(theta[1]+x[1])
    dxdt=[(r-u1-theta[3])*x[0],-(r*x[0])/theta[2]+u1*(u2-x[1])]
    
    return dxdt
    
#Correct model structure
  
def langmuir(x,t,U,theta):
    
    (tau,T)=U    #residence time and temperature
     
    (K1, K2, A3, Ea3, A4, Ea4, A5, Ea5, K6, K7, K8)=theta
    
    H_0=0.0013
    Pin=5
    H_T=H_0*np.exp(1700*((1/T)-(1/298.15)))
    kLa=50
    CLsat_O2=Pin*H_T
    
    (HMF, HFCA, FFCA, FDCA, O2)=x
    
    #Arrhenius
    #K3=A3*np.exp(-Ea3/(R*T))
    #K4=A4*np.exp(-Ea4/(R*T))
    #K5=A5*np.exp(-Ea5/(R*T))
    
    
    
    omega=(K2*O2)/(1+K1*HMF+K2*O2+HFCA/K6+FFCA/K7+FDCA/K8)**2
    
    rate=np.array([K1*K3*HMF*omega, (K4/K6)*HFCA*omega, (K5/K7)*FFCA*omega])
    
    #differential equations
    dxdt=np.zeros((len(x),1))
    
    dxdt=[-(length/tau)*rate[0],(length/tau)*(rate[0]-rate[1]),(length/tau)*(rate[1]-rate[2]),(length/tau)*rate[2],(length/tau)*(kLa*(CLsat_O2-O2)-0.5*(rate[0]+rate[1]+rate[2]))]
    return dxdt
    
    
def langmuir_ethanol(n,L,U,theta):
    
    A=np.array([16.5,-4.79])
    B=np.array([-9134.6,4386.0])
    
    K220=np.array([97.1,0.089,0.0011])
    
    Ea=np.array([36.25,12.95,1.6E-4])
    
    (BL_Et, BL_Ac, BL_EA, BL_H2)=(10.4,98.4,41.2,2.5E-4)
    
    stoichiometry=np.array([[-1,+1,0,+1],[-1,-1,+1,+1],[0,-2,0,0]])
    
    (Ptot,T)=U    #residence time and temperature
     
    (W,N2)=theta
    
    Tref=273.15+220.0
    
    (Et, Ac, EA, H2)=n
    
    #Arrhenius
    K1=K220[0]*np.exp((Ea[0]/R)*((1/Tref)-(1/T)))
    K2=K220[1]*np.exp((Ea[1]/R)*((1/Tref)-(1/T)))
    K3=K220[2]*np.exp((Ea[2]/R)*((1/Tref)-(1/T)))
       
    Keq1=np.exp(A[0]+B[0]/T)
    
    Keq2=np.exp(A[1]+B[1]/T)
    
    P_Et=Et*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_Ac=Ac*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_EA=EA*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    P_H2=H2*Ptot*0.986923/(Et+Ac+EA+H2+N2)
    
    rate=np.array([  K1*BL_Et*P_Et*(1-(1/Keq1)*(P_H2*P_Ac/P_Et))/((1+BL_Et*P_Et+BL_Ac*P_Ac+BL_H2*P_H2+BL_EA*P_EA)**2),
                     K2*BL_Et*BL_Ac*P_Et*P_Ac*(1-(1/Keq2)*(P_EA*P_H2/(P_Et*P_Ac)))/((1+BL_Et*P_Et+BL_Ac*P_Ac+BL_H2*P_H2+BL_EA*P_EA)**2),
                     K3*P_Ac**2])
    
    #differential equations
    
    
    
    dndt=np.array(W*np.dot(np.transpose(stoichiometry),rate))
    return dndt
    
    

#The objective function for the design of experiment shall depend on:
#the number of experimental points to be designed,
#constraints on the sampling frequency

#Correct model structure for case study on microbial population

def tametal(x,t,Ca,theta):
    
    alpha=1+theta[5]*(1-np.exp(-Ca[0]*t*theta[6]/100.0))
    
    G=theta[1]*(1-10**(x[0]-theta[4]))
    
    K=((theta[2]*Ca[0]**theta[3])/(Ca[0]**theta[3]+(alpha*theta[0])**theta[3]))
    
    dxdt=np.log10(np.exp(1.0))*(G-K)
    
    return dxdt

def tametal_incorrect(x,t,Ca,theta):
    
    alpha=1+theta[5]*(1-np.exp(-t*theta[6]/100.0))
    
    G=theta[1]*(1-10**(x[0]-theta[4]))
    
    K=((theta[2]*Ca[0]**theta[3])/(Ca[0]**theta[3]+(alpha*theta[0])**theta[3]))
    
    dxdt=np.log10(np.exp(1.0))*(G-K)
    
    return dxdt


#Function measurement generates a measurement in silico given initial values for the states, time-invariant inputs, sampling time and standard deviation of measurement error

def measurement(x0,U,t,sigma,theta,model,measurable):
    
    X=integrate.odeint(model,x0,[0.0,t],(U, theta,))
    #X=np.array([np.exp(X[1,0]),X[1,1],X[1,2],X[1,3],np.exp(X[1][-1])])
    measurement=np.abs(X[1]+np.random.normal(0.0,sigma))
    return [x0,U,t,measurement.tolist(),measurable]


#Function loglikelihood fits an incorrect model to the experimental data. It was preferred to integrate the model for every measurement singularly sacrifying computational efficiency

def loglikelihood(theta, dataset, sigma, model, measurable):
    
    residuals=np.zeros(shape=(len(dataset),len(sigma)))
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=[dataset[i][3] for i in range(0,len(dataset))]
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], theta))
        predictions[i]=X[1]
    
    residuals=((measurements-predictions)/sigma)**2
    
    objective=0
    
    for i in measurable:
        objective=objective+np.sum(residuals[:,i])
            
    return objective
      
      
#experiment performs a virtual experiment given initial states (list), time-invariant control input variables (list) and sampling times (list)  
  
def experiment(x0,U,st,sigma,theta,model,measurable):
    
    dataset=np.array([measurement(x0,U,st[0],sigma,theta,model,measurable)])
    for i in range(1,len(st)):
        dataset=np.append(dataset,[measurement(x0,U,st[i],sigma,theta,model,measurable)],axis=0)
        
    return dataset

#sensitivity evaluates the gradient of the measurable model output in the parameter space
def sensitivity(x0,U,t,theta,model,epsilon):
    #epsilon=0.0001
    sensitivity_matrix=np.ndarray(shape=(len(theta),len(x0)))
    perturbed_X=np.ndarray(shape=(len(theta),len(x0)))
    unperturbed_X=integrate.odeint(model,x0,[0.0,t],args=(U,theta,))[1]
    unperturbed_X=np.repeat([unperturbed_X],len(theta),axis=0)
    perturbed_theta=np.repeat([theta],len(theta),axis=0)+np.eye(len(theta))*epsilon
    for i in range(0,len(theta)):
        perturbed_X[i]=integrate.odeint(model,x0,[0.0,t],args=(U,perturbed_theta[i],))[1]
    
    sensitivity_matrix=np.subtract(perturbed_X,unperturbed_X)/epsilon
    
    return sensitivity_matrix
    
    
def relative_sensitivity(x0,U,t,theta,model,epsilon):
    #epsilon=0.0001
    sensitivity_matrix=np.ndarray(shape=(len(theta),len(x0)))
    perturbed_X=np.ndarray(shape=(len(theta),len(x0)))
    unperturbed_X=integrate.odeint(model,x0,[0.0,t],args=(U,theta,))[1]
    unperturbed_X=np.repeat([unperturbed_X],len(theta),axis=0)
    perturbed_theta=np.repeat([theta],len(theta),axis=0)+np.diag(theta)*epsilon
    for i in range(0,len(theta)):
        perturbed_X[i]=integrate.odeint(model,x0,[0.0,t],args=(U,perturbed_theta[i],))[1]
    
    sensitivity_matrix=np.divide(np.subtract(perturbed_X,unperturbed_X),np.transpose(epsilon*np.repeat([theta],len(x0),axis=0)))
    
    return sensitivity_matrix
            

#expected Fisher calculates the Fisher information matrix related to a single measurement point
def expected_fisher(x0,U,t,theta,sigma,model,measurable):
    epsilon=0.001
    Q=relative_sensitivity(x0,U,t,theta,model,epsilon)
    Fisher=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        Fisher=Fisher+(1/(sigma[i]**2))*np.outer(Q[:,i],Q[:,i])
        
    return Fisher
    
    

        
        
        
    
def expected_covariance(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable):
    
    U=design_vector
    #t=design_vector[-1]
    
    #V_theta=np.zeros(shape=(len(theta),len(theta)))
    
    F=expected_fisher(x0,U,length,theta,sigma,model,measurable)
    
    V_theta=np.linalg.inv(Obs_Fisher+F)
    
    return V_theta
    
def expected_covariance_novel(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model):
    
    U=design_vector[:-1]
    t=design_vector[-1]
    
    #V_theta=np.zeros(shape=(len(theta),len(theta)))
    
    F=expected_fisher(x0,U,t,theta,sigma,model,measurable)
    
    D=deviation(x0,U,t,theta,sigma,model,measurable,true_parameters,true_model)
    
    V_theta=np.linalg.inv(Obs_Fisher+F+D)
    
    return V_theta
    
    
def extended_fisher(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model):
    
    U=design_vector[:-1]
    t=design_vector[-1]
    
    #V_theta=np.zeros(shape=(len(theta),len(theta)))
    
    F=expected_fisher(x0,U,t,theta,sigma,model,measurable)
    
    D=deviation(x0,U,t,theta,sigma,model,measurable,true_parameters,true_model)
    
    ext_fisher=F+D
    
    return ext_fisher
    
def correlation(covariance):
    correlation=np.ndarray(shape=(len(covariance),len(covariance)))
    
    for i in range(0,len(covariance)):
        for j in range(0,len(covariance)):
            correlation[i,j]=covariance[i,j]/(np.sqrt(covariance[i,i]*covariance[j,j]))
    
    return correlation
    
    
def A_conventional_design(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable):
    
    V_theta=expected_covariance(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable)
    
    
    objective_function=np.trace(V_theta)
    
    return objective_function
    
def A_fisher_design(design_vector,x0,theta,sigma,model,measurable):
    
    U=design_vector[:-1]
    t=design_vector[-1]
    
    objective_function=np.trace(expected_fisher(x0,U,t,theta,sigma,model,measurable))
    
    return -objective_function
    
def D_conventional_design(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable):
    
    V_theta=expected_covariance(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable)
    
    
    objective_function=np.linalg.det(V_theta)
    
    return objective_function
    
    
def D_conventional_design_ethanol(design_vector,theta,sigma,model,Obs_Fisher,measurable):
    
    U=design_vector[0:2]
    
    initial=np.zeros(4)
    
    initial[0]=design_vector[-1]
    
    initial[1:4]=x0[1:4]
    
    V_theta=expected_covariance(U,initial,theta,sigma,model,Obs_Fisher,measurable)
    
    
    objective_function=np.linalg.det(V_theta)
    
    return np.log10(objective_function)
    

def A_novel_design(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model):
    
    V_theta=expected_covariance_novel(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model)
    
    
    objective_function=np.trace(V_theta)
    
    return objective_function
    
def D_novel_design(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model):
    
    V_theta=expected_covariance_novel(design_vector,x0,theta,sigma,model,Obs_Fisher,measurable,true_parameters,true_model)
    
    
    objective_function=np.linalg.det(V_theta)
    
    return objective_function
        
        
def preliminary_guess(ranges,dataset,sigma,model,measurable,num_guess):
    
    ranges=np.transpose(ranges)
    
    objective_function=np.zeros(num_guess)
    theta=np.zeros(shape=(num_guess,len(ranges[0])))
    
    for i in range(0,num_guess):
        theta[i]=ranges[0]+(ranges[1]-ranges[0])*np.random.random(size=len(ranges[0]))
        objective_function[i]=loglikelihood(theta[i], dataset, sigma, model, measurable)
        
    return theta[objective_function.tolist().index(min(objective_function))]
    
def model_output(theta,x0,t,U,model,variable):
    
    X=integrate.odeint(model,x0,[0.0,t],args=(U,theta,))[1]
    
    return X[variable]
        
def deviation(x0,U,t,theta,sigma,model,measurable,true_parameters,true_model):
    
    deviation=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        deviation=deviation+((1/(sigma[i]**2))*(model_output(theta,x0,t,U,model,i)-model_output(true_parameters,x0,t,U,true_model,i))*nd.Hessian(model_output)(theta,x0,t,U,model,i))
    
    return deviation
    
def observed_deviation(x0,U,t,theta,sigma,model,measurable,measurements):
    
    deviation=np.zeros(shape=(len(theta),len(theta)))
    
    for i in measurable:
        deviation=deviation+((1/(sigma[i]**2))*(model_output(theta,x0,t,U,model,i)-measurements[i])*nd.Hessian(model_output)(theta,x0,t,U,model,i))
    
    return deviation
    
def observed_fisher(dataset,sigma,theta,model):
    
    observed_information=np.zeros(shape=(len(dataset),2,len(theta),len(theta)))
    
    for i in range(0,len(dataset)):
        observed_information[i,0]=expected_fisher(dataset[i,0],dataset[i,1],dataset[i,2],theta,sigma,model,dataset[i,4])
        observed_information[i,1]=observed_deviation(dataset[i,0],dataset[i,1],dataset[i,2],theta,sigma,model,dataset[i,4],dataset[i,3])
    
    return observed_information

def excelsave(XX,YY,function,filename):
    
    wb=Workbook()

    ws=wb.active
    
    if XX.shape!=function.shape:
        function=function.reshape(XX.shape)
         
    
    for i in range(len(function)):
        for j in range(len(function)):
            ws.cell(row=i+2, column=1, value=(XX[i,0]/1))
            ws.cell(row=1, column=j+2, value=(YY[0,j]/1))
            ws.cell(row=i+2, column=j+2, value=function[i,j])
        
    wb.save(filename)
    
    return
    

    
def confidence_ellipsoid(total_covariance,significance,MLE,parameters):
    
    array_length=200
    covariance=np.zeros(shape=(2,2))
    
    for i in [0,1]:
        for j in [0,1]:
            covariance[i,j]=total_covariance[parameters[i],parameters[j]]
    
    eigenvalues,v=np.linalg.eig(covariance)
    
    if all([x>0 for x in eigenvalues]):
        eigenvalues=np.sqrt(eigenvalues)
    
        z=st.norm.ppf(significance)    
    
    
    
        ax = plt.subplot(111, aspect='equal')

        ell = Ellipse(xy=(MLE[parameters[0]],MLE[parameters[1]]),
                  width=eigenvalues[0]*z*2, height=eigenvalues[1]*z*2,
                  angle=np.rad2deg(np.arccos(v[0,0])))
        ell.set_facecolor('blue')
        ax.add_artist(ell)
    
    
        plt.show()
    
    
        X=np.linspace(-eigenvalues[0]*z,eigenvalues[0]*z,array_length)
    
        Y=np.zeros(shape=(2*len(X),2))
    
        for i in range(0,len(X)):
            Y[i,0]=X[i]
            Y[i+len(X),0]=X[-(i+1)]
            Y[i,1]=(eigenvalues[1]*z)*np.sqrt(1-((Y[i,0]**2)/((eigenvalues[0]*z)**2)))
            Y[i+len(X),1]=-(eigenvalues[1]*z)*np.sqrt(1-((Y[i+len(X),0]**2)/((eigenvalues[0]*z)**2)))
            Y[i]=np.dot(v,Y[i])+[MLE[parameters[0]],MLE[parameters[1]]]
            Y[i+len(X)]=np.dot(v,Y[i+len(X)])+[MLE[parameters[0]],MLE[parameters[1]]]
    
    else:
        Y=np.zeros(shape=(2*array_length,2))
    
    return Y

    
def export_array(array,filename):
    wb=Workbook()
    
    ws1=wb.active
    
    for i in range(0,len(array)):
        for j in range(0,len(array[i])):
            ws1.cell(row=1+i, column=1+j, value=array[i][j])
            
    wb.save(filename)
    
    return
              
                
def export_confidence_ellipsoids(observed_sensitivity,observed_deviation,significance,MLE,filename):
    
    wb=Workbook()
    
    ws1=wb.active
    
    ws1.title=('from_observed_sensitivity')
    
    index=1
    
    for i in range(0,len(MLE)):
        for j in range(i+1,len(MLE)):
            ell=confidence_ellipsoid(np.linalg.inv(observed_sensitivity),significance,MLE,[i,j])
            ws1.cell(row=1, column=index, value=(i+1))
            ws1.cell(row=1, column=index+1, value=(j+1))
            for k in range(len(ell)):
                ws1.cell(row=k+2, column=index, value=ell[k,0])
                ws1.cell(row=k+2, column=index+1, value=ell[k,1])
            index=index+3
    
    ws2=wb.create_sheet(title='from_observed_information')
    
    index=1
        
    for i in range(0,len(MLE)):
        for j in range(i+1,len(MLE)):
            ell=confidence_ellipsoid(np.linalg.inv(observed_sensitivity+observed_deviation),significance,MLE,[i,j])
            ws2.cell(row=1, column=index, value=(i+1))
            ws2.cell(row=1, column=index+1, value=(j+1))
            for k in range(len(ell)):
                ws2.cell(row=k+2, column=index, value=ell[k,0])
                ws2.cell(row=k+2, column=index+1, value=ell[k,1])
            index=index+3
            
    wb.save(filename)
    
    return
    
def det_deviation(design_vector,x0,theta,sigma,model,measurable,true_parameters,true_model):
    
    U=design_vector[:-1]
    t=design_vector[-1]
    
    objective_function=np.linalg.det(deviation(x0,U,t,theta,sigma,model,measurable,true_parameters,true_model))
    
    return -objective_function
    
    
def parity_plot(dataset, model, theta):
    
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=np.array([dataset[i][3] for i in range(0,len(dataset))])
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], theta))
        predictions[i]=X[1]
        
    plt.scatter(predictions,measurements)
    plt.show()
    
    return
    
    
def full_factorial(model, design_space, levels, sigma):
    
    #Generation of designs
    
    (F_range, P_range, T_range)=design_space
    
    (T_levels, P_levels, F_levels)=levels
    
    T,P,F=np.mgrid[T_range[0]:T_range[1]:np.complex(0,T_levels),P_range[0]:P_range[1]:np.complex(0,P_levels),F_range[0]:F_range[1]:np.complex(0,F_levels)]
    
    dataset=np.array(np.prod(levels))
    
    dataset=experiment([F.ravel()[0],0.000001,0.000001,N2],[P.ravel()[0],T.ravel()[0]],[length],sigma,[2.0,N2],langmuir_ethanol,[0,1,2,3])
    
    for i in range(1,np.prod(T.shape)):
        
        dataset=np.append(dataset,experiment([F.ravel()[i],0.000001,0.000001,N2],[P.ravel()[i],T.ravel()[i]],[length],sigma,[2.0,N2],langmuir_ethanol,[0,1,2,3]),axis=0)

    return dataset
    
    
def MBDM(theta, dataset, sigma, model, measurable, tolerance):
    
    residuals=np.zeros(shape=(len(dataset),len(sigma)))
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=[dataset[i][3] for i in range(0,len(dataset))]
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], theta))
        predictions[i]=X[1]
    
    residuals=((measurements-predictions)/sigma)**2
    
    residuals=residuals[:,measurable]
    
       
    objective=0
    
    switcher=np.ones(len(dataset))
    
    for i in range(0,len(dataset)):
        contribution=np.min([0.0, -len(measurable)*tolerance**2+np.sum(residuals[i,:])])
        objective=objective+contribution
        
    return objective
    
    
def MBDM_switcher(theta, dataset, sigma, model, measurable, tolerance):
    
    residuals=np.zeros(shape=(len(dataset),len(sigma)))
    predictions=np.zeros(shape=(len(dataset),len(sigma)))
    measurements=[dataset[i][3] for i in range(0,len(dataset))]
    
    for i in range(0,len(dataset)):
        X=integrate.odeint(model,dataset[i][0],[0.0,dataset[i][2]],(dataset[i][1], theta))
        predictions[i]=X[1]
    
    residuals=((measurements-predictions)/sigma)**2
    
    residuals=residuals[:,measurable]
    
       
    objective=0
    
    switcher=np.ones(len(dataset))
    
    for i in range(0,len(dataset)):
        contribution=np.min([0.0, -len(measurable)*tolerance**2+np.sum(residuals[i,:])])
        objective=objective+contribution
        if contribution==0.0:
            switcher[i]=0
        else:
            switcher[i]=1
    
    total_residual=np.sum(residuals[switcher==1,:])
    
    return total_residual, switcher
    
    
def exp_design(design_vector, model, parameters, observed_information, sigma, no_exp, criterion, output):
    
    normalised_conditions=design_vector.reshape((no_exp,3))
    
    conditions=np.add(np.transpose(design_space)[0],np.multiply(normalised_conditions,np.subtract(np.transpose(design_space)[1],np.transpose(design_space)[0])))
    
    
    F_hat=np.zeros(shape=(observed_information.shape))
    
    
    
    F_hat=sum(expected_fisher([conditions[i,0],0.0000001,0.0000001,4E-3],[conditions[i,1],conditions[i,2]],length,parameters,sigma, powerlaw_ethanol, measurable) for i in range(0,no_exp))
    
    if output=='objective':
    
        if criterion=='D':
        
            objective=np.log(np.linalg.det(np.linalg.inv(F_hat+observed_information)))
        
        elif criterion=='A':
        
            objective=np.trace(np.linalg.inv(F_hat+observed_information))
        
        elif criterion=='E':
        
            e,v=np.linalg.eig(np.linalg.inv(F_hat+observed_information))
            objective=np.max(e)
    
        return objective
        
    elif output=='covariance':
        
        return np.linalg.inv(F_hat+observed_information)
        
        
def t_test(parameters, covariance, significance, dof):
    
    variances=np.diag(covariance)
    
    t_ref=st.t.interval(1-(2*(1-significance)),dof)[1]
    
    t_values=np.array(parameters/(st.t.interval(significance,dof)[1]*np.sqrt(variances)))
    
    return t_values, t_ref, np.all(t_values>t_ref)
    
    
def predicted_information(conditions, theta, sigma, model):
    
    predicted_information=np.array([expected_fisher([conditions[i,0], 1e-06, 1e-06, 0.057],[conditions[i,1],conditions[i,2]],1.0,theta,sigma,model,measurable) for i in range(0,len(conditions))])

    return np.array([np.trace(predicted_information[i]) for i in range(0,len(predicted_information))])


def predicted_information_D(conditions, theta, sigma, model, observed_information):
    
    predicted_information=np.array([observed_information+expected_fisher([conditions[i,0], 1e-06, 1e-06, 0.057],[conditions[i,1],conditions[i,2]],1.0,theta,sigma,model,measurable) for i in range(0,len(conditions))])

    return np.array([np.linalg.det(predicted_information[i]) for i in range(0,len(predicted_information))])
        
    
def reliability_map(dataset, switchers, gamma):
    Y=2*(switchers-0.5*np.ones(len(switchers)))
    
    X_not_scaled=np.array([[dataset[i,0][0],dataset[i,1][0],dataset[i,1][1]] for i in range(0,len(dataset))])
    
    X=np.zeros((len(X_not_scaled),len(X_not_scaled[0])),dtype=np.float)

    for i in range(len(X_not_scaled)):
        for j in range(len(X_not_scaled[0])):
            X[i,j]=1*(X_not_scaled[i,j]-X_not_scaled[:,j].min())/(X_not_scaled[:,j].max()-X_not_scaled[:,j].min())

    #plane flowrate and pressure

    I=svm.SVC(C=1.0, kernel='rbf', gamma=gamma, class_weight='balanced')

    I.fit(X,Y)
    
    return I
    
def export_reliability(filename):
    
    XX,YY,ZZ=np.mgrid[0.0:1.0:200j,0.0:1.0:200j,0.0:1.0:200j]
    contours=np.zeros(shape=(200,200))
    mappa=np.zeros(shape=(200,200,200))
    '''
    for i,P in enumerate(np.linspace(0.0,1.0,200)):
        for j,F in enumerate(np.linspace(0.0,1.0,200)):
            for k,T in enumerate(np.linspace(0.0,1.0,200)):
    '''            
    mappa=clf.decision_function(np.c_[XX.ravel(),YY.ravel(),ZZ.ravel()]).reshape(XX.shape)
    
                
    for i,P in enumerate(np.linspace(0.0,1.0,200)):
        for j,F in enumerate(np.linspace(0.0,1.0,200)):
            try:
                contours[i,j]=np.linspace(0.0,1.0,200)[mappa[:,i,j].tolist().index(np.max([n for n in mappa[:,i,j] if n <0]))]
            except ValueError:
                #print 'contour lies outside the investigated window.'
                contours[i,j]=np.nan
    
    XX,YY=np.mgrid[0.0:1.0:200j,0.0:1.0:200j]
    
    excelsave(XX,YY,contours,filename)
    
    return
    
    
def export_reliability_2D(filename):
    
    XX,YY=np.mgrid[0.0:1.0:200j,0.0:1.0:200j]
    contour=np.zeros(shape=(200,200))
              
    contour=clf.decision_function(np.c_[XX.ravel(),YY.ravel()]).reshape(XX.shape)
    
    excelsave(XX,YY,contour,filename)
    
    return
    
def constraints(x):
    
    x_normalised=x.reshape(len(x)/3,3)
    
    
    constraint=np.min(clf.decision_function(x_normalised))
    
    return constraint
    
def initial_guess_design(no_exp):
    
    designed_exp=np.zeros(shape=(no_exp,3))
    for i in range(0,no_exp):
        acceptable=False
        while acceptable==False:
            designed_exp[i]=np.random.uniform(size=3)
            if constraints(designed_exp[i])>0:
                acceptable=True
                
    return designed_exp.ravel()
    
    